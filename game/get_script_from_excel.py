#coding=utf-8
from openpyxl import load_workbook
def raskdr():
    wb = load_workbook('xlsxs/Naomi раскадровка.xlsx')

    sheet = wb['N8']


    for i in range(2, 570):
        a = sheet['A'+str(i)]
        e = sheet['E'+str(i)]
        f = sheet['F'+str(i)]
        #try:
        if a.value is not None:
            print('    show sc i_' + a.value + ' with my_dissolve')

        if f.value and e.value:
            if f.font.i:
                print('    "'+e.value+'"'+' "{i}'+f.value+'{/i}"')
            else:
                print('    "'+e.value+'"'+' "'+f.value+'"')
        
        elif f.value:
            #if '"' in f.value:
            #    print("'{i}"+f.value+"{/i}'")

            #else:
            print('    "'+f.value+'"')

        #except:
        #    pass
raskdr()

# def ttable():
#     wb = load_workbook('Активности (1) (1).xlsx')

#     sheet = wb['QLogs']
#     rtrns = {}

#     for i in range(2, 300):

#         b = sheet['B'+str(i)]
#         c = sheet['C'+str(i)]
#         #try:
#         if b.value is not None:
#             if '_' in b.value:
#                 x = b.value.split('_')[0]
#                 if x not in rtrns:
#                     rtrns.update({x:{}})
#                 rtrns[x].update({int(b.value.split('_')[1]):int(c.value)})
#                 #print()

#     print(rtrns)

# ttable()


# def ttable():
#     wb = load_workbook('xlsxs/Активности (3).xlsx')
#     chars = {}
#     sheet = wb['Char Info']
#     for i in range(2, 500):
#         name        = sheet['A'+str(i)].value
#         surname     = sheet['B'+str(i)].value
        
#         if name is not None:
#             gender      = sheet['D'+str(i)].value
#             age         = sheet['C'+str(i)].value
#             description = sheet['E'+str(i)].value
#             if surname is None:
#                 chars.update({name.title():''})
                
#             else:
#                 chars.update({name.title():surname.title()})

#             #chars[name.title()].update({'gender':gender})
#             #chars[name.title()].update({'description':description})
#             #chars[name.title()].update({'age':age})
        



#     print(chars)


# ttable()